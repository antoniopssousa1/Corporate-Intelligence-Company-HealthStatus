"""
EXCEL DASHBOARD - Replica da Dashboard Streamlit em Excel
Cria uma dashboard interativa com formatação profissional
"""
import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import get_connection
from config import GOLD_DIR

# ============================================================================
# ESTILOS E CORES
# ============================================================================

# Cores do tema (consistentes com Streamlit)
COLORS = {
    'header_bg': '1E3A5F',      # Azul escuro
    'header_font': 'FFFFFF',    # Branco
    'excellent': '38EF7D',      # Verde
    'good': 'A8E063',           # Verde claro
    'fair': 'F2C94C',           # Amarelo
    'concerning': 'EB5757',     # Laranja/Vermelho
    'poor': 'C31432',           # Vermelho escuro
    'bronze': 'CD7F32',         # Bronze
    'silver': 'A9A9A9',         # Prata
    'gold': 'DAA520',           # Dourado
    'light_bg': 'F0F2F6',       # Cinza claro
    'border': 'CCCCCC',         # Cinza médio
}

# Estilos pré-definidos
HEADER_FONT = Font(bold=True, color=COLORS['header_font'], size=12)
HEADER_FILL = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
TITLE_FONT = Font(bold=True, color=COLORS['header_bg'], size=16)
SUBTITLE_FONT = Font(bold=True, color=COLORS['header_bg'], size=12)
THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border'])
)

def get_health_fill(status):
    """Retorna cor de fundo baseada no status de saúde"""
    color_map = {
        'Excellent': COLORS['excellent'],
        'Good': COLORS['good'],
        'Fair': COLORS['fair'],
        'Concerning': COLORS['concerning'],
        'Poor': COLORS['poor']
    }
    color = color_map.get(status, COLORS['light_bg'])
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def format_number(value, format_type='number'):
    """Formata números para exibição"""
    if pd.isna(value) or value is None:
        return "N/A"
    
    if format_type == 'currency_billions':
        if abs(value) >= 1e12:
            return f"${value/1e12:.2f}T"
        elif abs(value) >= 1e9:
            return f"${value/1e9:.2f}B"
        elif abs(value) >= 1e6:
            return f"${value/1e6:.2f}M"
        else:
            return f"${value:,.0f}"
    elif format_type == 'percentage':
        return f"{value*100:.1f}%"
    elif format_type == 'ratio':
        return f"{value:.2f}"
    else:
        return f"{value:,.0f}"


def create_excel_dashboard():
    """Cria a dashboard Excel completa"""
    print("=" * 60)
    print("EXCEL DASHBOARD - Criando Dashboard Interativa")
    print("=" * 60)
    
    conn = get_connection()
    
    # Carregar dados
    print("\n📊 Carregando dados...")
    
    df_kpi = pd.read_sql('''
        SELECT * FROM gold_kpi_dashboard 
        ORDER BY ticker, fiscal_year DESC
    ''', conn)
    
    df_health = pd.read_sql('''
        SELECT * FROM gold_financial_health 
        ORDER BY ticker, fiscal_year DESC
    ''', conn)
    
    df_companies = pd.read_sql('''
        SELECT ticker, company_name, sector FROM silver_companies 
        ORDER BY ticker
    ''', conn)
    
    df_income = pd.read_sql('''
        SELECT * FROM silver_income_statement 
        ORDER BY ticker, fiscal_year DESC
    ''', conn)
    
    df_balance = pd.read_sql('''
        SELECT * FROM silver_balance_sheet 
        ORDER BY ticker, fiscal_year DESC
    ''', conn)
    
    print(f"  ✓ {len(df_companies)} empresas carregadas")
    print(f"  ✓ {len(df_kpi)} registos KPI")
    print(f"  ✓ {len(df_health)} registos Health")
    
    # Criar workbook
    wb = Workbook()
    
    # ========================================================================
    # SHEET 1: DASHBOARD PRINCIPAL
    # ========================================================================
    print("\n📊 Criando Dashboard Principal...")
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    # Título principal
    ws_dash.merge_cells('B2:K2')
    ws_dash['B2'] = "📊 CORPORATE FINANCIAL HEALTH DASHBOARD"
    ws_dash['B2'].font = Font(bold=True, size=20, color=COLORS['header_bg'])
    ws_dash['B2'].alignment = Alignment(horizontal='center')
    
    ws_dash.merge_cells('B3:K3')
    ws_dash['B3'] = "Sistema de Análise de Saúde Financeira - Top 10 Tech Companies"
    ws_dash['B3'].font = Font(size=12, italic=True, color='666666')
    ws_dash['B3'].alignment = Alignment(horizontal='center')
    
    # ---- SECÇÃO: Ranking por Health Score ----
    ws_dash['B5'] = "🏆 RANKING POR HEALTH SCORE"
    ws_dash['B5'].font = TITLE_FONT
    
    # Obter dados mais recentes por empresa
    latest_year = df_health['fiscal_year'].max()
    df_latest = df_health[df_health['fiscal_year'] == latest_year].sort_values('health_score', ascending=False)
    
    # Cabeçalhos da tabela
    headers = ['#', 'Empresa', 'Ticker', 'Health Score', 'Status', 'Net Margin', 'ROE', 'D/E Ratio']
    for col, header in enumerate(headers, start=2):
        cell = ws_dash.cell(row=7, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Dados do ranking
    for idx, (_, row) in enumerate(df_latest.iterrows(), start=1):
        row_num = 7 + idx
        
        # Rank
        ws_dash.cell(row=row_num, column=2, value=idx).border = THIN_BORDER
        
        # Empresa
        ws_dash.cell(row=row_num, column=3, value=row['company_name']).border = THIN_BORDER
        
        # Ticker
        ws_dash.cell(row=row_num, column=4, value=row['ticker']).border = THIN_BORDER
        
        # Health Score
        cell_score = ws_dash.cell(row=row_num, column=5, value=row['health_score'])
        cell_score.border = THIN_BORDER
        cell_score.number_format = '0.0'
        cell_score.alignment = Alignment(horizontal='center')
        
        # Status (com cor)
        cell_status = ws_dash.cell(row=row_num, column=6, value=row['health_status'])
        cell_status.fill = get_health_fill(row['health_status'])
        cell_status.border = THIN_BORDER
        cell_status.alignment = Alignment(horizontal='center')
        cell_status.font = Font(bold=True, color='FFFFFF' if row['health_status'] in ['Excellent', 'Poor', 'Concerning'] else '000000')
        
        # Net Margin
        cell_nm = ws_dash.cell(row=row_num, column=7, value=row['net_margin'])
        cell_nm.number_format = '0.0%'
        cell_nm.border = THIN_BORDER
        
        # ROE
        cell_roe = ws_dash.cell(row=row_num, column=8, value=row['roe'])
        cell_roe.number_format = '0.0%'
        cell_roe.border = THIN_BORDER
        
        # Debt to Equity
        cell_de = ws_dash.cell(row=row_num, column=9, value=row['debt_to_equity'])
        cell_de.number_format = '0.00'
        cell_de.border = THIN_BORDER
    
    # ---- SECÇÃO: KPIs Agregados ----
    ws_dash['L5'] = "📈 KPIs AGREGADOS (Último Ano)"
    ws_dash['L5'].font = TITLE_FONT
    
    # Calcular métricas agregadas
    total_revenue = df_latest.merge(df_kpi[df_kpi['fiscal_year'] == latest_year], on='ticker')['revenue'].sum()
    total_income = df_latest.merge(df_kpi[df_kpi['fiscal_year'] == latest_year], on='ticker')['net_income'].sum()
    avg_health = df_latest['health_score'].mean()
    
    kpi_data = [
        ('💰 Receita Total', format_number(total_revenue, 'currency_billions')),
        ('📊 Lucro Líquido Total', format_number(total_income, 'currency_billions')),
        ('🏥 Health Score Médio', f"{avg_health:.1f}"),
        ('🌟 Empresas Excellent', str(len(df_latest[df_latest['health_status'] == 'Excellent']))),
        ('✅ Empresas Good', str(len(df_latest[df_latest['health_status'] == 'Good']))),
        ('⚠️ Empresas Fair/Concern', str(len(df_latest[df_latest['health_status'].isin(['Fair', 'Concerning'])]))),
    ]
    
    for idx, (label, value) in enumerate(kpi_data):
        row_num = 7 + idx
        cell_label = ws_dash.cell(row=row_num, column=12, value=label)
        cell_label.font = Font(bold=True)
        cell_label.fill = PatternFill(start_color=COLORS['light_bg'], end_color=COLORS['light_bg'], fill_type='solid')
        cell_label.border = THIN_BORDER
        
        cell_value = ws_dash.cell(row=row_num, column=13, value=value)
        cell_value.font = Font(bold=True, size=14, color=COLORS['header_bg'])
        cell_value.alignment = Alignment(horizontal='center')
        cell_value.border = THIN_BORDER
    
    # Ajustar larguras das colunas
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 5
    ws_dash.column_dimensions['C'].width = 25
    ws_dash.column_dimensions['D'].width = 10
    ws_dash.column_dimensions['E'].width = 12
    ws_dash.column_dimensions['F'].width = 12
    ws_dash.column_dimensions['G'].width = 12
    ws_dash.column_dimensions['H'].width = 10
    ws_dash.column_dimensions['I'].width = 10
    ws_dash.column_dimensions['L'].width = 22
    ws_dash.column_dimensions['M'].width = 15
    
    # ---- GRÁFICO: Health Score Bar Chart ----
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "Health Score por Empresa"
    chart1.y_axis.title = "Health Score"
    
    # Dados para o gráfico (colunas E = Health Score, linha 8 a 17)
    data = Reference(ws_dash, min_col=5, min_row=7, max_row=7+len(df_latest), max_col=5)
    cats = Reference(ws_dash, min_col=4, min_row=8, max_row=7+len(df_latest))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    chart1.width = 15
    chart1.height = 10
    
    ws_dash.add_chart(chart1, "B20")
    
    # ========================================================================
    # SHEET 2: ANÁLISE DETALHADA POR EMPRESA
    # ========================================================================
    print("📊 Criando Análise Detalhada...")
    ws_detail = wb.create_sheet("Análise Detalhada")
    
    # Título
    ws_detail.merge_cells('B2:M2')
    ws_detail['B2'] = "📋 ANÁLISE DETALHADA POR EMPRESA"
    ws_detail['B2'].font = TITLE_FONT
    ws_detail['B2'].alignment = Alignment(horizontal='center')
    
    # Criar dropdown de seleção de empresa
    ws_detail['B4'] = "Selecione a Empresa:"
    ws_detail['B4'].font = Font(bold=True)
    
    # Lista de empresas para dropdown
    companies_list = df_companies['ticker'].tolist()
    dv = DataValidation(type="list", formula1=f'"{",".join(companies_list)}"', allow_blank=False)
    dv.error = "Selecione uma empresa da lista"
    dv.errorTitle = "Empresa Inválida"
    ws_detail.add_data_validation(dv)
    
    ws_detail['C4'] = companies_list[0]  # Default: primeira empresa
    dv.add(ws_detail['C4'])
    ws_detail['C4'].fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    ws_detail['C4'].border = THIN_BORDER
    
    # Tabela de dados de todas as empresas (para VLOOKUP/INDEX-MATCH)
    # Escrever dados em área oculta
    start_row = 50
    ws_detail.cell(row=start_row, column=2, value="DATA_TABLE")
    
    headers_detail = ['Ticker', 'Company', 'Year', 'Revenue', 'Net Income', 'Health Score', 
                      'Status', 'Current Ratio', 'Quick Ratio', 'Gross Margin', 
                      'Net Margin', 'ROE', 'ROA', 'D/E', 'D/A', 'FCF Margin']
    
    for col, header in enumerate(headers_detail, start=2):
        ws_detail.cell(row=start_row+1, column=col, value=header)
    
    # Merge dos dados
    df_merged = df_health.merge(df_kpi[['ticker', 'fiscal_year', 'revenue', 'net_income']], 
                                  on=['ticker', 'fiscal_year'], how='left')
    
    for idx, (_, row) in enumerate(df_merged.iterrows()):
        r = start_row + 2 + idx
        ws_detail.cell(row=r, column=2, value=row['ticker'])
        ws_detail.cell(row=r, column=3, value=row['company_name'])
        ws_detail.cell(row=r, column=4, value=row['fiscal_year'])
        ws_detail.cell(row=r, column=5, value=row['revenue'])
        ws_detail.cell(row=r, column=6, value=row['net_income'])
        ws_detail.cell(row=r, column=7, value=row['health_score'])
        ws_detail.cell(row=r, column=8, value=row['health_status'])
        ws_detail.cell(row=r, column=9, value=row['current_ratio'])
        ws_detail.cell(row=r, column=10, value=row['quick_ratio'])
        ws_detail.cell(row=r, column=11, value=row['gross_margin'])
        ws_detail.cell(row=r, column=12, value=row['net_margin'])
        ws_detail.cell(row=r, column=13, value=row['roe'])
        ws_detail.cell(row=r, column=14, value=row['roa'])
        ws_detail.cell(row=r, column=15, value=row['debt_to_equity'])
        ws_detail.cell(row=r, column=16, value=row['debt_to_assets'])
        ws_detail.cell(row=r, column=17, value=row['free_cash_flow_margin'])
    
    # Fórmulas para exibir dados da empresa selecionada
    ws_detail['B6'] = "🏢 Empresa:"
    ws_detail['C6'] = f'=VLOOKUP($C$4,$B${start_row+2}:$Q${start_row+2+len(df_merged)},2,FALSE)'
    
    # Secção: Métricas Principais
    ws_detail['B8'] = "💰 MÉTRICAS PRINCIPAIS"
    ws_detail['B8'].font = SUBTITLE_FONT
    
    metrics_main = [
        ('Health Score', 7, '0.0'),
        ('Status', 8, '@'),
        ('Revenue', 5, '$#,##0'),
        ('Net Income', 6, '$#,##0'),
    ]
    
    for idx, (label, col_idx, fmt) in enumerate(metrics_main):
        row = 9 + idx
        ws_detail.cell(row=row, column=2, value=label).font = Font(bold=True)
        cell = ws_detail.cell(row=row, column=3)
        cell.value = f'=VLOOKUP($C$4,$B${start_row+2}:$Q${start_row+2+len(df_merged)},{col_idx},FALSE)'
        cell.number_format = fmt
        cell.border = THIN_BORDER
    
    # Secção: Rácios de Liquidez
    ws_detail['B14'] = "💧 LIQUIDEZ"
    ws_detail['B14'].font = SUBTITLE_FONT
    
    liquidity_ratios = [
        ('Current Ratio', 9, '0.00', 1.5),
        ('Quick Ratio', 10, '0.00', 1.0),
    ]
    
    row = 15
    for label, col_idx, fmt, benchmark in liquidity_ratios:
        ws_detail.cell(row=row, column=2, value=label).font = Font(bold=True)
        cell = ws_detail.cell(row=row, column=3)
        cell.value = f'=VLOOKUP($C$4,$B${start_row+2}:$Q${start_row+2+len(df_merged)},{col_idx},FALSE)'
        cell.number_format = fmt
        cell.border = THIN_BORDER
        
        ws_detail.cell(row=row, column=4, value=f"Benchmark: ≥{benchmark}").font = Font(size=9, italic=True, color='666666')
        row += 1
    
    # Secção: Rentabilidade
    ws_detail['B18'] = "💰 RENTABILIDADE"
    ws_detail['B18'].font = SUBTITLE_FONT
    
    profit_ratios = [
        ('Gross Margin', 11, '0.0%', 0.30),
        ('Net Margin', 12, '0.0%', 0.10),
        ('ROE', 13, '0.0%', 0.15),
        ('ROA', 14, '0.0%', 0.05),
    ]
    
    row = 19
    for label, col_idx, fmt, benchmark in profit_ratios:
        ws_detail.cell(row=row, column=2, value=label).font = Font(bold=True)
        cell = ws_detail.cell(row=row, column=3)
        cell.value = f'=VLOOKUP($C$4,$B${start_row+2}:$Q${start_row+2+len(df_merged)},{col_idx},FALSE)'
        cell.number_format = fmt
        cell.border = THIN_BORDER
        
        ws_detail.cell(row=row, column=4, value=f"Benchmark: ≥{benchmark*100:.0f}%").font = Font(size=9, italic=True, color='666666')
        row += 1
    
    # Secção: Alavancagem
    ws_detail['B24'] = "⚖️ ALAVANCAGEM"
    ws_detail['B24'].font = SUBTITLE_FONT
    
    leverage_ratios = [
        ('Debt-to-Equity', 15, '0.00', '≤1.0'),
        ('Debt-to-Assets', 16, '0.0%', '≤50%'),
    ]
    
    row = 25
    for label, col_idx, fmt, benchmark in leverage_ratios:
        ws_detail.cell(row=row, column=2, value=label).font = Font(bold=True)
        cell = ws_detail.cell(row=row, column=3)
        cell.value = f'=VLOOKUP($C$4,$B${start_row+2}:$Q${start_row+2+len(df_merged)},{col_idx},FALSE)'
        cell.number_format = fmt
        cell.border = THIN_BORDER
        
        ws_detail.cell(row=row, column=4, value=f"Benchmark: {benchmark}").font = Font(size=9, italic=True, color='666666')
        row += 1
    
    # Ajustar larguras
    ws_detail.column_dimensions['B'].width = 18
    ws_detail.column_dimensions['C'].width = 20
    ws_detail.column_dimensions['D'].width = 18
    
    # ========================================================================
    # SHEET 3: COMPARAÇÃO DE EMPRESAS
    # ========================================================================
    print("📊 Criando Comparação de Empresas...")
    ws_compare = wb.create_sheet("Comparação")
    
    # Título
    ws_compare.merge_cells('B2:L2')
    ws_compare['B2'] = "📊 COMPARAÇÃO ENTRE EMPRESAS"
    ws_compare['B2'].font = TITLE_FONT
    ws_compare['B2'].alignment = Alignment(horizontal='center')
    
    # Tabela comparativa
    compare_headers = ['Ticker', 'Empresa', 'Health Score', 'Status', 'Revenue ($B)', 
                       'Net Margin', 'ROE', 'D/E Ratio', 'FCF Margin']
    
    for col, header in enumerate(compare_headers, start=2):
        cell = ws_compare.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    # Dados
    df_compare = df_latest.merge(df_kpi[df_kpi['fiscal_year'] == latest_year][['ticker', 'revenue', 'net_income']], 
                                  on='ticker', how='left')
    
    for idx, (_, row) in enumerate(df_compare.iterrows()):
        r = 5 + idx
        
        ws_compare.cell(row=r, column=2, value=row['ticker']).border = THIN_BORDER
        ws_compare.cell(row=r, column=3, value=row['company_name']).border = THIN_BORDER
        
        cell_score = ws_compare.cell(row=r, column=4, value=row['health_score'])
        cell_score.number_format = '0.0'
        cell_score.border = THIN_BORDER
        
        cell_status = ws_compare.cell(row=r, column=5, value=row['health_status'])
        cell_status.fill = get_health_fill(row['health_status'])
        cell_status.border = THIN_BORDER
        cell_status.alignment = Alignment(horizontal='center')
        
        cell_rev = ws_compare.cell(row=r, column=6, value=row['revenue']/1e9 if pd.notna(row['revenue']) else None)
        cell_rev.number_format = '0.00'
        cell_rev.border = THIN_BORDER
        
        cell_nm = ws_compare.cell(row=r, column=7, value=row['net_margin'])
        cell_nm.number_format = '0.0%'
        cell_nm.border = THIN_BORDER
        
        cell_roe = ws_compare.cell(row=r, column=8, value=row['roe'])
        cell_roe.number_format = '0.0%'
        cell_roe.border = THIN_BORDER
        
        cell_de = ws_compare.cell(row=r, column=9, value=row['debt_to_equity'])
        cell_de.number_format = '0.00'
        cell_de.border = THIN_BORDER
        
        cell_fcf = ws_compare.cell(row=r, column=10, value=row['free_cash_flow_margin'])
        cell_fcf.number_format = '0.0%'
        cell_fcf.border = THIN_BORDER
    
    # Adicionar formatação condicional para Health Score (Data Bars)
    ws_compare.conditional_formatting.add(
        f'D5:D{4+len(df_compare)}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                    color="63C384", showValue=True, minLength=None, maxLength=None)
    )
    
    # Ajustar larguras
    ws_compare.column_dimensions['B'].width = 10
    ws_compare.column_dimensions['C'].width = 25
    ws_compare.column_dimensions['D'].width = 14
    ws_compare.column_dimensions['E'].width = 12
    ws_compare.column_dimensions['F'].width = 14
    ws_compare.column_dimensions['G'].width = 12
    ws_compare.column_dimensions['H'].width = 10
    ws_compare.column_dimensions['I'].width = 10
    ws_compare.column_dimensions['J'].width = 12
    
    # ========================================================================
    # SHEET 4: TENDÊNCIAS HISTÓRICAS
    # ========================================================================
    print("📊 Criando Tendências Históricas...")
    ws_trends = wb.create_sheet("Tendências")
    
    # Título
    ws_trends.merge_cells('B2:L2')
    ws_trends['B2'] = "📈 TENDÊNCIAS HISTÓRICAS"
    ws_trends['B2'].font = TITLE_FONT
    ws_trends['B2'].alignment = Alignment(horizontal='center')
    
    # Tabela de evolução de Health Score por ano
    years = sorted(df_health['fiscal_year'].unique(), reverse=True)[:5]
    
    headers_trends = ['Ticker'] + [str(y) for y in years] + ['Δ 5Y']
    for col, header in enumerate(headers_trends, start=2):
        cell = ws_trends.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    for idx, ticker in enumerate(df_companies['ticker'].unique()):
        r = 5 + idx
        ws_trends.cell(row=r, column=2, value=ticker).border = THIN_BORDER
        
        ticker_data = df_health[df_health['ticker'] == ticker].set_index('fiscal_year')
        
        first_score = None
        last_score = None
        
        for col_idx, year in enumerate(years):
            cell = ws_trends.cell(row=r, column=3+col_idx)
            if year in ticker_data.index:
                score = ticker_data.loc[year, 'health_score']
                cell.value = score
                cell.number_format = '0.0'
                
                if first_score is None:
                    first_score = score
                last_score = score
            else:
                cell.value = "-"
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Variação 5 anos
        cell_delta = ws_trends.cell(row=r, column=3+len(years))
        if first_score and last_score:
            delta = first_score - last_score  # Mais recente - mais antigo
            cell_delta.value = delta
            cell_delta.number_format = '+0.0;-0.0;0.0'
            if delta > 0:
                cell_delta.font = Font(color='008000', bold=True)
            elif delta < 0:
                cell_delta.font = Font(color='FF0000', bold=True)
        cell_delta.border = THIN_BORDER
        cell_delta.alignment = Alignment(horizontal='center')
    
    # Ajustar larguras
    ws_trends.column_dimensions['B'].width = 10
    for col in range(3, 3+len(years)+1):
        ws_trends.column_dimensions[get_column_letter(col)].width = 10
    
    # ========================================================================
    # SHEET 5: DADOS BRUTOS
    # ========================================================================
    print("📊 Adicionando Dados Brutos...")
    ws_data = wb.create_sheet("Dados")
    
    # Exportar df_health completo
    for r_idx, row in enumerate(dataframe_to_rows(df_health, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
    
    # ========================================================================
    # GUARDAR FICHEIRO
    # ========================================================================
    output_dir = os.path.join(GOLD_DIR, 'excel_export')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'DASHBOARD_INTERATIVA.xlsx')
    
    wb.save(output_path)
    print(f"\n✅ Dashboard Excel criada com sucesso!")
    print(f"   📁 Localização: {output_path}")
    print(f"\n📋 Sheets criadas:")
    print(f"   1. Dashboard - Visão geral e ranking")
    print(f"   2. Análise Detalhada - Métricas por empresa (com dropdown)")
    print(f"   3. Comparação - Tabela comparativa")
    print(f"   4. Tendências - Evolução histórica")
    print(f"   5. Dados - Dados brutos para análise")
    
    return output_path


if __name__ == "__main__":
    create_excel_dashboard()
